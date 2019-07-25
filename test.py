import feedparser
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from smtplib import SMTP_SSL
from urllib.parse import quote
from openpyxl import Workbook
import ssl
import sys

print(sys.argv)

news_rss_url = 'http://rss.etnews.com/Section903.xml'
economy_rss_url = 'http://rss.etnews.com/02.xml'
tech_rss_url = 'http://rss.etnews.com/04.xml'

SMTP_SERVER = "smtp.naver.com"
SMTP_PORT = "465"
SMTP_USER = "pjhyl1127"
SMTP_PASSWORD = "zxczxc1127"

def send_mail(name, addr, contents, attachment=False) :
    msg = MIMEMultipart("alternative")

    if attachment :
        msg = MIMEMultipart('mixed')
    
    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name+'님, 오늘의 전자신문 기사입니다'

    text = MIMEText(contents)
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octet-stream')
        f = open(attachment, 'rb')
        file_contents = f.read()
        file_data.set_payload(file_contents)
        encoders.encode_base64(file_data)

        from os.path import basename
        filename = basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment', filename=filename)
        msg.attach(file_data)

    smtp = SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()

xlsx = Workbook()
std = xlsx.get_sheet_by_name('Sheet')
xlsx.remove_sheet(std)

def popular_news() :
    news_sheet = xlsx.create_sheet('인기기사')
    news_sheet.append(['기사 제목', '링크', '날짜'])
    ssl._create_default_https_context = ssl._create_unverified_context
    news_list = feedparser.parse(news_rss_url)
    for news in news_list['items'] :
        news_sheet.append([news['title'], news['link'], news['published']])

def economy_news() :
    economy_sheet = xlsx.create_sheet('경제')
    economy_sheet.append(['기사 제목', '링크', '날짜'])
    news_list = feedparser.parse(economy_rss_url)
    for news in news_list['items'] :
        economy_sheet.append([news['title'], news['link'], news['published']])

def tech_news() :
    tech_sheet = xlsx.create_sheet('기술')
    tech_sheet.append(['기사 제목', '링크', '날짜'])
    news_list = feedparser.parse(tech_rss_url)
    for news in news_list['items'] :
        tech_sheet.append([news['title'], news['link'], news['published']])

keyword = sys.argv

if(keyword == '경제'):
    economy_news()
elif(keyword == '기술'):
    tech_news()
else:
    popular_news()
    
file_name = 'news_list.xlsx'
xlsx.save(file_name)
send_mail('Park', 'pjhyl1127@gmail.com', '뉴스 수집 결과입니다.', file_name)
